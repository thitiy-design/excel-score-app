"""
Flask Backend – ระบบตรวจสอบคะแนนพนักงาน

Score ใน Excel มีสองรูปแบบ:
  1. ข้อความ "5/5", "5 / 5", "10/10"  → อ่านค่าได้โดยตรง
  2. ตัวเลขล้วน เช่น 3, 5  + Excel number_format เช่น '0" / 4"' → ต้องอ่าน format
"""

import os, io, re
import pandas as pd
import openpyxl
from flask import Flask, request, jsonify, send_file, send_from_directory
from flask_cors import CORS

BASE_DIR     = os.path.dirname(os.path.abspath(__file__))
FRONTEND_DIR = os.path.join(BASE_DIR, '..', 'frontend')

app = Flask(__name__, static_folder=None)
CORS(app)
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024

ALLOWED_EXTENSIONS = {'xlsx'}
REQUIRED_COLUMNS   = {'Timestamp', 'Score', 'รหัสพนักงาน', 'ชื่อ-สกุล'}


def allowed_file(filename: str) -> bool:
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def extract_denominator_from_format(fmt: str):
    """
    ดึง denominator จาก Excel number_format
    เช่น '0" / 4"' → 4,  '0" / 5"' → 5,  'General' → None
    """
    if not fmt or fmt in ('General', '@', ''):
        return None
    m = re.search(r'/\s*"?\s*(\d+)', fmt)
    if m:
        return int(m.group(1))
    m = re.search(r'(\d+)\s*"?\s*$', fmt)
    if m:
        return int(m.group(1))
    return None


def parse_score_string(val):
    """
    แปลง string Score เป็น (numerator, denominator)
    รองรับ: "5/5", "5 / 5", "10/10", "4/4"
    คืน None ถ้าแปลงไม่ได้
    """
    s = str(val).strip()
    m = re.match(r'^(\d+(?:\.\d+)?)\s*/\s*(\d+(?:\.\d+)?)$', s)
    if m:
        return (float(m.group(1)), float(m.group(2)))
    return None


def read_excel_with_scores(file_bytes: bytes):
    """
    อ่านไฟล์ Excel ด้วย openpyxl เพื่อดึงทั้ง value และ number_format ของคอลัมน์ Score
    คืน (df, score_col_index) โดย df มีคอลัมน์ '_score_num' และ '_score_den'
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active

    # หา header row
    headers = [cell.value for cell in ws[1]]
    headers_stripped = [str(h).strip() if h is not None else '' for h in headers]

    score_col_idx = None
    for i, h in enumerate(headers_stripped):
        if h == 'Score':
            score_col_idx = i
            break

    if score_col_idx is None:
        return None, None

    # อ่านข้อมูลทุกแถว
    rows_data = []
    for row in ws.iter_rows(min_row=2):
        row_dict = {}
        for i, cell in enumerate(row):
            col_name = headers_stripped[i] if i < len(headers_stripped) else f'col_{i}'
            row_dict[col_name] = cell.value

            # ถ้าเป็นคอลัมน์ Score ให้ดึง number_format ด้วย
            if i == score_col_idx:
                fmt = cell.number_format
                den_from_fmt = extract_denominator_from_format(fmt)
                row_dict['_score_format'] = fmt
                row_dict['_den_from_fmt'] = den_from_fmt

        rows_data.append(row_dict)

    df = pd.DataFrame(rows_data)
    if not df.empty:
        df.columns = df.columns.str.strip()

    return df, score_col_idx


def resolve_score(row) -> tuple:
    """
    หา (numerator, denominator) จากแถวข้อมูล
    ลำดับความสำคัญ:
      1. ถ้า Score เป็น string รูปแบบ X/Y → ใช้ค่านั้น
      2. ถ้า Score เป็นตัวเลข + มี _den_from_fmt → (score, den_from_fmt)
      3. อื่นๆ → None
    """
    val = row.get('Score')
    den_fmt = row.get('_den_from_fmt')

    if pd.isna(val) if not isinstance(val, str) else val.strip() == '':
        return None

    # ลอง parse string X/Y ก่อน
    if isinstance(val, str):
        parsed = parse_score_string(val)
        if parsed:
            return parsed

    # ถ้าเป็นตัวเลข
    try:
        num = float(val)
        if den_fmt is not None:
            return (num, float(den_fmt))
        # ตัวเลขล้วนไม่มี format — ไม่รู้ denominator
        return None
    except (ValueError, TypeError):
        return None


def is_perfect(row) -> bool:
    score = resolve_score(row)
    if score is None:
        return False
    numerator, denominator = score
    return numerator == denominator


def fmt_score_display(row) -> str:
    score = resolve_score(row)
    if score is None:
        return str(row.get('Score', '')).strip()
    n = int(score[0]) if score[0] == int(score[0]) else score[0]
    d = int(score[1]) if score[1] == int(score[1]) else score[1]
    return f'{n}/{d}'


def process_excel(file_bytes: bytes) -> dict:
    try:
        df, score_col_idx = read_excel_with_scores(file_bytes)

        if df is None:
            return {'success': False, 'error': 'ไม่พบคอลัมน์ Score ในไฟล์'}

        # ตรวจสอบคอลัมน์จำเป็น
        missing = REQUIRED_COLUMNS - set(df.columns)
        if missing:
            return {
                'success': False,
                'error': f'ไม่พบคอลัมน์: {", ".join(sorted(missing))}',
                'found_columns': [c for c in df.columns if not c.startswith('_')]
            }

        total_rows = len(df)

        # กรองแถวที่ได้คะแนนเต็ม
        mask = df.apply(is_perfect, axis=1)
        perfect_df = df[mask].copy()

        # เพิ่มคอลัมน์ Score_display
        perfect_df['Score_display'] = perfect_df.apply(fmt_score_display, axis=1)

        # หา max_scores ที่พบในไฟล์
        max_scores = set()
        for _, row in df.iterrows():
            score = resolve_score(row)
            if score and score[1] is not None:
                max_scores.add(int(score[1]))

        # แปลง Timestamp
        perfect_df['Timestamp'] = perfect_df['Timestamp'].apply(
            lambda v: str(v) if v is not None else ''
        ).str.replace('NaT', '').str.replace('None', '')

        # แปลง รหัสพนักงาน
        perfect_df['รหัสพนักงาน'] = perfect_df['รหัสพนักงาน'].apply(
            lambda v: str(int(float(v))) if v is not None and str(v).replace('.', '').isdigit() else str(v or '')
        )

        perfect_df['ชื่อ-สกุล'] = perfect_df['ชื่อ-สกุล'].fillna('').astype(str)

        records = perfect_df[['Timestamp', 'รหัสพนักงาน', 'ชื่อ-สกุล', 'Score_display']].rename(
            columns={'Score_display': 'Score'}
        ).reset_index(drop=True).to_dict(orient='records')

        return {
            'success':       True,
            'total_rows':    total_rows,
            'perfect_count': len(records),
            'max_scores':    sorted(max_scores),
            'data':          records
        }

    except Exception as e:
        import traceback
        return {'success': False, 'error': f'เกิดข้อผิดพลาด: {str(e)}', 'detail': traceback.format_exc()}


# ── Serve Frontend ───────────────────────────────────────────────

@app.route('/')
def index():
    return send_from_directory(FRONTEND_DIR, 'index.html')

@app.route('/static/<path:filename>')
def static_files(filename):
    return send_from_directory(os.path.join(FRONTEND_DIR, 'static'), filename)


# ── API Endpoints ────────────────────────────────────────────────

@app.route('/api/health')
def health():
    return jsonify({'status': 'ok'})


@app.route('/api/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': 'ไม่พบไฟล์ในคำขอ'}), 400
    file = request.files['file']
    if not file.filename:
        return jsonify({'success': False, 'error': 'ยังไม่ได้เลือกไฟล์'}), 400
    if not allowed_file(file.filename):
        return jsonify({'success': False, 'error': 'รองรับเฉพาะไฟล์ .xlsx เท่านั้น'}), 400

    result = process_excel(file.read())
    return jsonify(result), 200 if result['success'] else 422


@app.route('/api/export', methods=['POST'])
def export_excel():
    try:
        body    = request.get_json()
        data    = (body or {}).get('data', [])
        columns = (body or {}).get('columns', None)
        if not data:
            return jsonify({'error': 'ไม่มีข้อมูลสำหรับ Export'}), 400

        df = pd.DataFrame(data)
        if columns:
            # ใช้เฉพาะ columns ที่มีจริงใน df
            cols = [c for c in columns if c in df.columns]
            df = df[cols]

        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Perfect Score')
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
            ws = writer.sheets['Perfect Score']
            for col_num in range(1, len(df.columns) + 1):
                cell = ws.cell(row=1, column=col_num)
                cell.font      = Font(bold=True, color='FFFFFF')
                cell.fill      = PatternFill('solid', start_color='2563EB')
                cell.alignment = Alignment(horizontal='center')
                ws.column_dimensions[get_column_letter(col_num)].width = 24

        output.seek(0)
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='perfect_score_employees.xlsx'
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500


if __name__ == '__main__':
    print("✅  Frontend : http://127.0.0.1:5000")
    print("✅  API Base : http://127.0.0.1:5000/api")
    app.run(debug=True, host='0.0.0.0', port=5000)
